"""excel_read / excel_write — port of src/tools/excel.ts (openpyxl instead of SheetJS)."""

import csv
import re
from pathlib import Path

import openpyxl
from langchain_core.tools import tool

from ..config import SCRATCH_DIR

MAX_ROWS = 100


def _read_xlsx(path: Path) -> dict[str, list[dict]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: dict[str, list[dict]] = {}
    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            out[ws.title] = []
            continue
        keys = [str(h) if h is not None else f"col{i}" for i, h in enumerate(header)]
        rows = []
        for row in rows_iter:
            rows.append({k: v for k, v in zip(keys, row)})
            if len(rows) >= MAX_ROWS:
                break
        out[ws.title] = rows
    wb.close()
    return out


def _read_csv(path: Path) -> dict[str, list[dict]]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = []
        for row in reader:
            rows.append(dict(row))
            if len(rows) >= MAX_ROWS:
                break
    return {"Sheet1": rows}


@tool
def excel_read(file_path: str) -> dict:
    """Read an .xlsx or .csv file from a local path. Returns sheet names and rows as JSON (capped at 100 rows per sheet)."""
    p = Path(file_path)
    if not p.exists():
        return {"success": False, "error": f"File not found: {file_path}"}
    if p.suffix.lower() not in (".xlsx", ".csv"):
        return {"success": False, "error": "Only .xlsx/.csv supported in v1"}
    try:
        data = _read_csv(p) if p.suffix.lower() == ".csv" else _read_xlsx(p)
        return {"success": True, "output": {"sheets": list(data.keys()), "data": data}}
    except Exception as e:
        return {"success": False, "error": str(e)}


@tool
def excel_write(filename: str, rows: list[dict], sheet_name: str = "Sheet1") -> dict:
    """Create an .xlsx file from JSON rows. Keys of the row objects become column headers. Returns the created file path."""
    try:
        out_dir = SCRATCH_DIR / "outbound"
        out_dir.mkdir(parents=True, exist_ok=True)
        safe = re.sub(r"\.\w+$", "", re.sub(r"[^\w.\-]", "_", filename)) + ".xlsx"
        out_path = out_dir / safe

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        if rows:
            keys = list(rows[0].keys())
            ws.append(keys)
            for row in rows:
                ws.append([row.get(k) for k in keys])
        wb.save(out_path)
        return {"success": True, "output": {"filePath": str(out_path)}}
    except Exception as e:
        return {"success": False, "error": str(e)}
